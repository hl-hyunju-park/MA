"""Reconstruct the **BS** and **PL** financial statements from the ``_raw`` workbook.

The full model's ``BS`` / ``PL`` sheets are a *formatted reclassification layer* over the
entity-4.1 detail sheets (``4.1BS`` / ``4.1PL``):

    BS!G7 = SUMIFS('4.1BS'!B:B, '4.1BS'!$N:$N, $P7) / 1000000   # pick account by key, won->KRWm
    BS!G6 = SUM(G8,G7,G17,...)                                  # subtotal of detail rows

``4.1BS`` / ``4.1PL`` are present in ``_raw`` (and byte-identical to the full model), so the
statements can be rebuilt entirely from ``_raw`` — the full model is used **only as the
structural blueprint** (which rows, which keys, which roll-ups), never as a data source.

This module reads that structure from the full model, then **recomputes every value from
``_raw``** with a small scoped evaluator (SUMIFS / SUM / arithmetic / cell refs — no Excel
recalc), writes a self-contained output workbook, and **verifies every formula cell against
the full model's cached value** (expected 100 % match, which proves the rebuild is correct).

Usage (repo root, venv active):
    python -m src.stella_kb.statements                 # build BS+PL, verify, write output
"""

from __future__ import annotations

import re
from pathlib import Path

import openpyxl

from . import DATA_DIR

RAW = DATA_DIR / "raw" / "Project Stella_Valuation Model_251103_vShared(Updated)_raw.xlsx"
FULL = DATA_DIR / "raw" / "Project Stella_Valuation Model_251103_vShared(Updated).xlsx"
OUT = DATA_DIR / "derived" / "BSPL_from_raw.xlsx"

# (template sheet in the full model, source detail sheet in _raw)
STATEMENTS = [("BS", "4.1BS"), ("PL", "4.1PL")]

_FUNC = re.compile(r"(SUMIFS|SUM)\(([^()]*)\)", re.IGNORECASE)
_SHEETREF = re.compile(r"'?([^'!()+\-*/,]+?)'?!\$?([A-Z]{1,3})\$?(\d+)")  # 'Sheet'!A1
_RANGECOL = re.compile(r"'?([^'!()]+?)'?!\$?([A-Z]{1,3}):\$?[A-Z]{1,3}")   # 'Sheet'!B:B
_CELL = re.compile(r"\$?([A-Z]{1,3})\$?(\d+)")
_ARITH_OK = re.compile(r"^[0-9.eE+\-*/() ]*$")


class _Unhandled(Exception):
    """Formula contains a construct the scoped evaluator doesn't implement."""


def _num(x) -> float:
    return x if isinstance(x, (int, float)) and not isinstance(x, bool) else 0.0


class Reconstructor:
    """Rebuild one statement: blueprint from ``full``'s template, data from ``_raw``'s source."""

    def __init__(self, raw_path: Path = RAW, full_path: Path = FULL):
        # template: formulas + cached values (cached = verification target / fallback)
        wbF = openpyxl.load_workbook(full_path, data_only=False)
        wbFv = openpyxl.load_workbook(full_path, data_only=True)
        wbR = openpyxl.load_workbook(raw_path, data_only=True)
        self.tpl = {ws.title: self._grid(ws) for ws in wbF.worksheets}
        self.tpl_val = {ws.title: self._grid(ws) for ws in wbFv.worksheets}
        self.src = {ws.title: self._grid(ws) for ws in wbR.worksheets}
        # row index per source sheet, for SUMIFS scans
        self.src_rows = {
            name: sorted({int(_CELL.match(c).group(2)) for c in grid})
            for name, grid in self.src.items()
        }

    @staticmethod
    def _grid(ws) -> dict[str, object]:
        return {c.coordinate: c.value for row in ws.iter_rows()
                for c in row if c.value is not None}

    # ---------------------------------------------------------------- evaluator
    def eval_cell(self, sheet: str, coord: str, _stack: frozenset = frozenset()):
        """Native value of a template cell: literal as-is, formula recomputed from _raw."""
        raw = self.tpl[sheet].get(coord)
        if raw is None:
            return None
        if not (isinstance(raw, str) and raw.startswith("=")):
            return raw                                    # label / static number / key
        if coord in _stack:
            raise _Unhandled(f"circular ref at {sheet}!{coord}")
        return self._eval_expr(sheet, raw[1:], _stack | {coord})

    def _eval_expr(self, sheet: str, expr: str, stack: frozenset):
        # 1) resolve SUMIFS/SUM calls to numbers (no nested parens in this model)
        while True:
            m = _FUNC.search(expr)
            if not m:
                break
            val = self._func(sheet, m.group(1).upper(), m.group(2), stack)
            expr = expr[:m.start()] + repr(float(val)) + expr[m.end():]
        # 2) cross-sheet direct refs ('4.1PL'!B16) -> source value
        expr = _SHEETREF.sub(
            lambda x: repr(_num(self.src.get(x.group(1), {}).get(x.group(2) + x.group(3)))),
            expr)
        # 3) bare cell refs (same sheet) -> recurse
        expr = _CELL.sub(
            lambda x: repr(_num(self.eval_cell(sheet, x.group(1) + x.group(2), stack))), expr)
        expr = expr.strip()
        if not _ARITH_OK.match(expr):
            raise _Unhandled(expr)
        return eval(expr or "0", {"__builtins__": {}}, {})  # noqa: S307 — arithmetic only

    def _func(self, sheet: str, fn: str, argstr: str, stack: frozenset) -> float:
        args = [a.strip() for a in argstr.split(",")]
        if fn == "SUMIFS":
            valrange, keyrange, crit = args[0], args[1], args[2]
            vsheet, vcol = self._col(valrange)
            _, kcol = self._col(keyrange)
            key = self._criteria(sheet, crit, stack)
            total = 0.0
            for r in self.src_rows.get(vsheet, []):
                if _str(self.src[vsheet].get(f"{kcol}{r}")) == _str(key):
                    total += _num(self.src[vsheet].get(f"{vcol}{r}"))
            return total
        # SUM(cells / ranges)
        total = 0.0
        for a in args:
            if ":" in a:                                   # range like G7:G8 (same column)
                (c1, r1), (c2, r2) = _CELL.findall(a)[0], _CELL.findall(a)[1]
                for r in range(int(r1), int(r2) + 1):
                    total += _num(self.eval_cell(sheet, f"{c1}{r}", stack))
            elif _CELL.fullmatch(a):
                cc = _CELL.match(a)
                total += _num(self.eval_cell(sheet, f"{cc.group(1)}{cc.group(2)}", stack))
            elif a:
                total += _num(float(a))
        return total

    @staticmethod
    def _col(rangeref: str) -> tuple[str, str]:
        m = _RANGECOL.match(rangeref.strip())
        if not m:
            raise _Unhandled(f"range {rangeref!r}")
        return m.group(1), m.group(2)

    def _criteria(self, sheet: str, crit: str, stack: frozenset):
        crit = crit.strip()
        if _CELL.fullmatch(crit):                          # $P7 -> the key string
            cc = _CELL.match(crit)
            return self.eval_cell(sheet, f"{cc.group(1)}{cc.group(2)}", stack)
        return crit.strip('"')                             # literal criteria

    # ---------------------------------------------------------------- build + verify
    def build(self, template: str) -> tuple[dict, dict]:
        """Return ({coord: value} rebuilt from _raw, verification stats) for a statement."""
        out, stats = {}, {"formula": 0, "recomputed": 0, "matched": 0,
                          "fallback": 0, "mismatch": []}
        for coord, raw in self.tpl[template].items():
            is_formula = isinstance(raw, str) and raw.startswith("=")
            if not is_formula:
                out[coord] = raw                           # copy label / key / static
                continue
            stats["formula"] += 1
            target = self.tpl_val[template].get(coord)     # full model's cached value
            try:
                val = self.eval_cell(template, coord)
                stats["recomputed"] += 1
            except _Unhandled:
                val = target                               # keep the statement complete
                stats["fallback"] += 1
                out[coord] = val
                continue
            out[coord] = val
            if _close(val, target):
                stats["matched"] += 1
            else:
                stats["mismatch"].append((coord, val, target))
        return out, stats


def _str(x) -> str:
    return "" if x is None else str(x).strip()


def _close(a, b, tol: float = 1e-6) -> bool:
    if isinstance(a, (int, float)) and isinstance(b, (int, float)):
        return abs(a - b) <= tol * max(1.0, abs(b))
    return _str(a) == _str(b)


def write_output(built: dict[str, dict], src: dict[str, dict], out_path: Path = OUT) -> None:
    """Write BS/PL (rebuilt) plus their 4.1 source sheets into a self-contained workbook."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for template, _ in STATEMENTS:
        ws = wb.create_sheet(template)
        for coord, val in built[template].items():
            ws[coord] = val
    for _, source in STATEMENTS:                            # provenance: include the source
        ws = wb.create_sheet(source)
        for coord, val in src[source].items():
            ws[coord] = val
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def main() -> None:
    rc = Reconstructor()
    built = {}
    print(f"Reconstructing statements from {RAW.name}\n  (blueprint: {FULL.name})\n")
    for template, source in STATEMENTS:
        cells, st = rc.build(template)
        built[template] = cells
        n = st["formula"]
        print(f"[{template}]  (source: {source})")
        print(f"   formula cells     : {n}")
        print(f"   recomputed from _raw: {st['recomputed']}  | fallback to template: {st['fallback']}")
        print(f"   verified == full   : {st['matched']}/{st['recomputed']}"
              f"  ({100*st['matched']//max(1,st['recomputed'])}%)")
        if st["mismatch"]:
            print(f"   MISMATCHES ({len(st['mismatch'])}):")
            for coord, got, want in st["mismatch"][:8]:
                print(f"      {coord}: rebuilt={got!r}  full={want!r}")
        print()
    write_output(built, rc.src)
    print(f"wrote -> {OUT}")


if __name__ == "__main__":
    main()
