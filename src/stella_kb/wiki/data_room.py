"""Ingest a nested **document data room** (the v0.3 corpus) into a queryable wiki.

v0.1/v0.2 are one canonical *formula model* + a few FDD decks; the v0.3 corpus is a real M&A
due-diligence data room — a deep folder tree of ~1.5k mixed-format files (the KDB Life target). The
formula-model pipeline (``dump_md → parse_llm → compile → index``) is single-workbook and coupled to
the cross-sheet **formula DAG** (``build_sheet_dag``/``write_ledgers``), which has no meaning across
hundreds of independent ledgers. So this module assembles a wiki for the data room from the *reused
pieces* of that pipeline, skipping the DAG machinery, and writes a **standard** ``index.json`` +
``pages/*.md`` the agent serves unchanged.

Pipeline (workbook-free):
  1. **curate**   — walk the corpus, drop the bulk/boilerplate sets (``curate.yaml`` exclude/include),
                    so vision/LLM spend only touches the DD-relevant core. Deterministic + offline.
  2. **(convert)**— ``convert.py`` first normalizes legacy formats in place (doc/docx/hwp/pptx/img →
                    pdf, xls → xlsx); run it before this build (``scripts/run_ingest_v03.sh``) or pass
                    ``--convert``. Here we just consume the resulting ``.xlsx`` / ``.pdf``.
  3. **spreadsheets → md grids** — ``dump_md.sheet_to_md`` per sheet (full cell fidelity, NO LLM);
                    aliases = the grid's text cells with their cell refs (provenance kept).
  4. **PDFs → vision pages**      — ``pdf_pages.build_pages``/``build_document`` per file, namespaced
                    by data-room section. The only LLM/vision spend.
  5. **assemble** — merge both streams into one index (``merge_into_index`` + ``dedup_alias_index``),
                    render ``INDEX.md``. Register the dataset in ``config.yaml`` to serve it.

    python -m src.stella_kb.wiki.data_room --plan            # curation dry-run + counts (offline)
    python -m src.stella_kb.wiki.data_room --only 3.         # build, but only the '3. 계리' section (pilot)
    python -m src.stella_kb.wiki.data_room --convert         # convert legacy formats first, then build
    python -m src.stella_kb.wiki.data_room                   # full curated build into MNA_WIKI_DATA
"""
from __future__ import annotations

import argparse
import json
import re
import sys
import unicodedata
from collections import Counter
from fnmatch import fnmatch
from pathlib import Path

import openpyxl

from ..config import (alias_stopwords, convert_root, curate_yaml, wiki_data_dir,
                      wiki_index_json, wiki_index_md, wiki_pages_dir)
from openpyxl.utils import get_column_letter

from . import pdf_pages as pp
from .dedup import dedup_alias_index
from .dump_md import _fmt
from .index import render_md

# A data-room asset register can be 10k+ rows / a 15 MB workbook. Stream it read_only (bounded
# memory) and cap the dumped grid so one pathological ledger can't produce a multi-MB page or OOM.
ROW_CAP = 2000
COL_CAP = 80
ALIAS_CAP = 200
# Vision is per-page; this room has 16 PDFs of 100–180 pages (tax forms, accounting policies, property
# registers) whose tails are boilerplate. Cap vision to the first N pages of every PDF — front matter,
# summaries and the substantive schedules live up front. Override with env MNA_PDF_PAGE_CAP.
PDF_PAGE_CAP = 25

# Built-in curation, mirrored in knowledge/v0.3/curate.yaml (the committed source of truth). Each
# pattern is matched (fnmatch) against a file's POSIX path *relative to the corpus root*. As of
# 2026-06-30 this is a **full ingest** — only the two genuinely unopenable files are excluded; the
# former boilerplate bulk-drops (영업계약/평가보고서/세무/정관·사규) were removed so the whole DD room is
# served. The cheaper-trim globs are documented in curate.yaml for easy restore.
DEFAULT_EXCLUDE: list[str] = [
    # Security-locked (NOT boilerplate): can't be opened, so excluded to keep rebuilds clean. Other
    # vintages of each series did ingest, so the data isn't lost.
    "*유형자산, 투자부동산 연도별 세부명세서 (FY2023.4분기).xlsx",   # password-encrypted OOXML (AES)
    "*FY2022_법인세 세무조정계산서.pdf",                          # DRM-wrapped ("<DOCUMENT SAFER")
]

def _nfc(s: str) -> str:
    """Normalize to NFC. This data room comes from macOS, whose filesystem stores Korean filenames
    in **NFD** (decomposed jamo) — visually identical to NFC but a different byte/codepoint sequence,
    so an NFD page key/alias would never match a user's NFC query. Normalizing every path-derived
    string (sections, page names, aliases, curation globs, ``--only``) keeps the wiki all-NFC."""
    return unicodedata.normalize("NFC", s)


SPREADSHEET_EXTS = {".xlsx", ".xlsm"}     # .xls is converted to .xlsx by convert.py first
# Legacy doc/image formats convert.py turns into .pdf; here we ingest the converted .pdf sibling.
CONVERT_TO_PDF = {".doc", ".docx", ".hwp", ".pptx", ".ppt", ".jpg", ".jpeg", ".tif", ".tiff"}
SKIP_NAMES = {".DS_Store"}


# --------------------------------------------------------------------------------------------------
# 1. curation
# --------------------------------------------------------------------------------------------------
def load_policy() -> tuple[list[str], list[str]]:
    """Read ``curate.yaml`` → ``(exclude, include)`` globs. Absent/empty → built-in
    ``DEFAULT_EXCLUDE`` and no include override. ``include`` re-admits files an ``exclude`` pattern
    would drop (evaluated after exclude), e.g. keep ``정관 20250331.pdf`` out of the bylaws bulk."""
    path = curate_yaml()
    if not path.exists():
        return list(DEFAULT_EXCLUDE), []
    try:
        import yaml
        data = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
    except Exception:  # noqa: BLE001 — a malformed curation file must not silently widen scope
        return list(DEFAULT_EXCLUDE), []
    exclude = data.get("exclude")
    include = data.get("include") or []
    return (list(exclude) if exclude else list(DEFAULT_EXCLUDE)), list(include)


def curate(root: Path, exclude: list[str], include: list[str]) -> list[Path]:
    """Walk ``root`` and return the curated file list (sorted, deterministic): every file except the
    ones an ``exclude`` glob drops, plus any ``include`` glob re-admits. Litter (``.DS_Store``) and
    non-files are skipped."""
    exclude = [_nfc(pat) for pat in exclude]
    include = [_nfc(pat) for pat in include]
    kept: list[Path] = []
    for p in sorted(root.rglob("*")):
        if not p.is_file() or p.name in SKIP_NAMES:
            continue
        rel = _nfc(p.relative_to(root).as_posix())   # disk paths are NFD on macOS — compare as NFC
        dropped = any(fnmatch(rel, pat) for pat in exclude)
        if dropped and not any(fnmatch(rel, pat) for pat in include):
            continue
        kept.append(p)
    return kept


def resolve_grains(files: list[Path]) -> tuple[list[Path], list[Path], list[Path]]:
    """Split the curated files into the two ingestable grains, following each legacy file to its
    converted sibling (produced by ``convert.py`` first): spreadsheets (``.xlsx``) and PDFs
    (native or converted). Returns ``(spreadsheets, pdfs, unconverted)`` — ``unconverted`` are legacy
    files whose converted sibling is missing (run ``convert`` first), reported so nothing is lost."""
    sheets: list[Path] = []
    pdfs: list[Path] = []
    unconverted: list[Path] = []
    seen_pdf: set[Path] = set()

    def add_pdf(p: Path) -> None:
        if p not in seen_pdf:
            seen_pdf.add(p)
            pdfs.append(p)

    for f in files:
        ext = f.suffix.lower()
        if ext in SPREADSHEET_EXTS:
            sheets.append(f)
        elif ext == ".xls":
            x = f.with_suffix(".xlsx")            # openpyxl can't read raw .xls — need convert.py's sibling
            (sheets if x.exists() else unconverted).append(x if x.exists() else f)
        elif ext == ".pdf":
            add_pdf(f)
        elif ext in CONVERT_TO_PDF:
            sib = f.with_suffix(".pdf")
            if sib.exists():
                add_pdf(sib)
            else:
                unconverted.append(f)
        # anything else (e.g. a stray .txt) is ignored
    return sheets, pdfs, unconverted


# --------------------------------------------------------------------------------------------------
# 2. namespacing — keep page keys unique + readable across a deep tree
# --------------------------------------------------------------------------------------------------
def section_of(path: Path, root: Path) -> str:
    """Top-level data-room folder = the wiki ``section`` (ToC bucket), e.g. ``2. 재무`` (NFC)."""
    rel = path.relative_to(root)
    return _nfc(rel.parts[0]) if len(rel.parts) > 1 else "(root)"


def _group_of(path: Path, root: Path) -> str:
    """Immediate parent folder = the wiki ``group`` (sub-bucket), e.g. ``2.6.3. 수익증권`` (NFC)."""
    rel = path.relative_to(root)
    return _nfc(rel.parts[-2]) if len(rel.parts) >= 2 else section_of(path, root)


def _uniquify(name: str, used: set[str]) -> str:
    """Make ``name`` unique within ``used`` by appending ``#2``, ``#3``, … and record it."""
    out = name
    n = 2
    while out in used:
        out = f"{name} #{n}"
        n += 1
    used.add(out)
    return out


def sheet_page_name(path: Path, sheet: str, root: Path, n_sheets: int, used: set[str]) -> str:
    """Namespaced, collision-free page key for one spreadsheet sheet:
    ``<parent-folder>__<file-stem>`` (+ ``__<sheet>`` only when the file has >1 sheet). '/' is
    escaped like the rest of the pipeline so key == file stem == ``[[wikilink]]`` target."""
    base = f"{_group_of(path, root)}__{_nfc(path.stem)}"
    if n_sheets > 1:
        base = f"{base}__{_nfc(sheet)}"
    return _uniquify(base.replace("/", "_"), used)


def pdf_doc_token(path: Path, root: Path, used: set[str]) -> str:
    """Per-PDF ``doc`` namespace for ``build_pages`` (prefixes ``FDD{n}`` page names). Folder +
    stem, kept unique so two reports never collide on an ``FDD3`` page key."""
    token = f"{_group_of(path, root)} / {_nfc(path.stem)}".replace("/", "_")
    return _uniquify(token, used)


# --------------------------------------------------------------------------------------------------
# 3. spreadsheet stream — dumped grids + cell-ref aliases (no LLM)
# --------------------------------------------------------------------------------------------------
def dump_sheet(name: str, ws) -> tuple[str, list[tuple[str, str]]]:
    """Stream one read_only worksheet → ``(markdown_grid, [(term, cellref), …])`` in a SINGLE pass
    (no random ``.cell()`` access — that's what OOMs on a 15 MB register). The grid is the cached
    values (formulas dropped: a data-room ledger isn't a formula model); text cells double as
    ``(term, cellref)`` aliases with exact provenance. Capped at ``ROW_CAP``×``COL_CAP`` with an
    overflow note so one giant sheet can't blow up the page."""
    rows: list[tuple[int, list[object]]] = []
    aliases: list[tuple[str, str]] = []
    seen: set[str] = set()
    max_c = 0
    row_overflow = False
    for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if r_idx > ROW_CAP:
            row_overflow = True
            break
        last = max((i for i, v in enumerate(row) if v is not None), default=-1)
        if last < 0:
            continue                                   # blank row — drop (keep the page compact)
        cells = list(row[:last + 1])
        max_c = max(max_c, len(cells))
        rows.append((r_idx, cells))
        for c_idx, v in enumerate(cells, start=1):
            if c_idx > COL_CAP or not isinstance(v, str) or len(aliases) >= ALIAS_CAP:
                continue
            t = re.sub(r"\s+", " ", v).strip(" *`")
            if not (2 <= len(t) <= 60) or pp._NUMERIC.match(t):
                continue
            k = pp._norm(t)
            if k and k not in seen:
                seen.add(k)
                aliases.append((t, f"{get_column_letter(c_idx)}{r_idx}"))

    cols = min(max_c, COL_CAP)
    col_overflow = max_c > COL_CAP
    lines = [f"# {name}", ""]
    if not rows:
        return "\n".join(lines) + "\n_(empty sheet)_\n", aliases
    # No grid-summary line ("N non-empty rows × M cols") — it's redundant boilerplate the LLM reads as
    # content. Keep only a truncation warning, and only when the sheet was actually clipped.
    if row_overflow or col_overflow:
        lines.append(f"> ⚠️ {ROW_CAP}×{COL_CAP} 초과 — 일부 잘림 (sheet has more)")
    lines += ["## Values", ""]
    letters = [get_column_letter(c) for c in range(1, cols + 1)]
    lines += ["| | " + " | ".join(letters) + " |", "|---|" + "---|" * cols]
    for r_idx, cells in rows:
        out = [_fmt(v) for v in cells[:cols]]
        out += [""] * (cols - len(out))
        lines.append(f"| **{r_idx}** | " + " | ".join(out) + " |")
    return "\n".join(lines) + "\n", aliases


def dump_spreadsheets(sheets: list[Path], root: Path, pages_dir: Path,
                      used_names: set[str]) -> tuple[dict, dict, dict]:
    """Dump every curated spreadsheet to md-grid wiki pages + the index pieces. Returns
    ``(entries, alias_add, tree_add)`` shaped exactly like ``pdf_pages.build_pages`` so the two
    streams merge through the same ``merge_into_index``."""
    pages_dir.mkdir(parents=True, exist_ok=True)
    entries: dict[str, dict] = {}
    aliases: dict[str, list] = {}
    tree: dict[str, dict] = {}

    for xlsx in sheets:
        try:                                  # read_only + data_only: stream cached values, low memory
            wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
        except Exception as e:  # noqa: BLE001 — a corrupt/encrypted/locked workbook shouldn't kill the build
            hint = " (password-protected or corrupt)" if "not a zip" in str(e).lower() else ""
            print(f"  !! skip (open failed){hint}: {xlsx.name}  ({e})")
            continue
        section, group, stem = section_of(xlsx, root), _group_of(xlsx, root), _nfc(xlsx.stem)
        names = wb.sheetnames
        for sheet_raw in names:
            ws = wb[sheet_raw]
            if not hasattr(ws, "iter_rows"):   # Chartsheet (chart-only tab) — no cell grid, skip
                continue
            sheet = _nfc(sheet_raw)
            name = sheet_page_name(xlsx, sheet_raw, root, len(names), used_names)
            md, grid_aliases = dump_sheet(name, ws)
            if not grid_aliases and md.endswith("_(empty sheet)_\n"):
                used_names.discard(name)       # reclaim the name; an empty sheet carries nothing
                continue
            (pages_dir / f"{name}.md").write_text(md, encoding="utf-8")
            terms = [t for t, _ in grid_aliases]
            page_aliases = [group, stem, sheet] + terms
            entries[name] = {
                "sheet": name,
                "title": f"{stem} · {sheet}",
                "desc": f"{section} / {group} — {stem} 시트 '{sheet}' (그리드 원문).",
                "section": section,
                "group": group,
                "kind": "ledger",
                "case": None,
                "unit": "",
                "period": None,
                "n_items": len(terms),
                "has_page": True,
                "aliases": page_aliases,
                "items": [{"label": t, "cell": ref, "role": "grid"} for t, ref in grid_aliases],
                "depends_on": [],
                "feeds_into": [],
                "source": "XLSX",
            }
            for term, ref in grid_aliases:
                aliases.setdefault(pp._norm(term), []).append(
                    {"page": name, "cell": ref, "term": term})
            for term in (group, stem, sheet):           # folder/file/sheet as routable aliases too (NFC)
                aliases.setdefault(pp._norm(term), []).append(
                    {"page": name, "cell": "", "term": term})
            tree.setdefault(section, {}).setdefault(group, []).append(name)
        wb.close()
    return entries, aliases, tree


# --------------------------------------------------------------------------------------------------
# 4. PDF stream — reuse pdf_pages, folder-namespaced
# --------------------------------------------------------------------------------------------------
def ingest_pdfs(pdfs: list[Path], root: Path, pages_dir: Path, used_docs: set[str],
                max_pages: int | None = PDF_PAGE_CAP) -> tuple[dict, dict, dict, dict]:
    """Run each PDF through the existing vision pipeline (``build_pages``), namespaced by its
    data-room section, capped at ``max_pages`` pages each. Returns merged
    ``(entries, alias_add, tree_add, documents)``.

    PDFs are **independent** in the data-room build (no cross-PDF formula DAG, and ``build_pages`` is
    called without an ``index``, so there are no cross-PDF xrefs), so several run **concurrently**
    (``pdf_file_concurrency``) — without overlap, a deep room drains one short PDF at a time and the
    shared vLLM batch sits half-empty. Doc tokens are pre-assigned *sequentially* (uniqueness needs
    deterministic order) and the per-PDF results are merged *in input order*, so the build stays
    byte-deterministic regardless of completion order."""
    from concurrent.futures import ThreadPoolExecutor

    pages_dir.mkdir(parents=True, exist_ok=True)
    # Curated first layer (decks.yaml), keyed by the per-PDF ``doc`` token. Same precedence as the
    # formula-model build: a deck block pins ``title``/``description`` (skips the LLM doc node) and
    # ``pages: {<FDD#>: "<name>"}`` freezes routed page names so a rebuild can't orphan routes.yaml.
    # Most v0.3 PDFs have no block → pure-LLM, unchanged; the file is purely additive.
    decks = pp._load_decks()
    if decks:
        print(f"  curated deck overrides for {len(decks)} doc(s)")

    # Pre-assign each PDF's doc token sequentially (``_uniquify`` mutates ``used_docs`` — not
    # thread-safe), so the concurrent builds below never race on the namespace.
    jobs = [(pdf, pdf_doc_token(pdf, root, used_docs)) for pdf in pdfs]

    def _build_one(job: tuple[Path, str]):
        pdf, doc = job
        deck = decks.get(doc) or {}
        title_pins = {k: v for k, v in (deck.get("pages") or {}).items()}
        print(f"  pdf: {pdf.relative_to(root)}  (doc={doc}, cap={max_pages})")
        try:
            e, a, t = pp.build_pages(str(pdf), pages_dir, doc=doc,
                                     title_pins=title_pins or None, max_pages=max_pages)
        except Exception as ex:  # noqa: BLE001 — one bad PDF shouldn't sink the whole room
            print(f"  !! skip (vision failed): {pdf.name}  ({ex})")
            return None
        doc_node = pp.build_document(doc, e, curated=deck or None) if e else None
        return doc, e, a, t, doc_node

    from ..config import pdf_file_concurrency
    workers = max(1, pdf_file_concurrency())
    if workers > 1 and len(jobs) > 1:
        with ThreadPoolExecutor(max_workers=workers) as ex:
            results = list(ex.map(_build_one, jobs))   # map preserves input order → deterministic merge
    else:
        results = [_build_one(j) for j in jobs]

    entries: dict[str, dict] = {}
    aliases: dict[str, list] = {}
    tree: dict[str, dict] = {}
    documents: dict[str, dict] = {}
    for res in results:
        if not res:
            continue
        doc, e, a, t, doc_node = res
        entries.update(e)
        for k, bucket in a.items():
            aliases.setdefault(k, []).extend(bucket)
        for sec, groups in t.items():
            dst = tree.setdefault(sec, {})
            for g, ns in groups.items():
                dst.setdefault(g, []).extend(ns)
        if doc_node:
            documents[doc] = doc_node
    return entries, aliases, tree, documents


# --------------------------------------------------------------------------------------------------
# 5. assemble
# --------------------------------------------------------------------------------------------------
def build(root: Path, *, only: str | None = None, no_pdf: bool = False) -> dict:
    """Curate → dump spreadsheets → ingest PDFs → assemble one wiki index. Writes
    ``<wiki_data>/wiki/{index.json, INDEX.md, pages/*.md}`` and returns the index dict."""
    exclude, include = load_policy()
    files = curate(root, exclude, include)
    if only:
        only = _nfc(only)
        files = [f for f in files if _nfc(f.relative_to(root).as_posix()).startswith(only)
                 or section_of(f, root).startswith(only)]
    sheets, pdfs, unconverted = resolve_grains(files)
    if unconverted:
        print(f"  !! {len(unconverted)} legacy file(s) not yet converted (run convert first) — "
              f"e.g. {unconverted[0].name}")
    print(f"==> curated {len(files)} file(s): {len(sheets)} spreadsheet(s), {len(pdfs)} pdf(s)"
          + (f"  [--only {only}]" if only else ""))

    pages_dir = wiki_pages_dir()
    used_names: set[str] = set()
    s_entries, s_aliases, s_tree = dump_spreadsheets(sheets, root, pages_dir, used_names)
    print(f"  dumped {len(s_entries)} spreadsheet page(s)")

    index = {"tree": {}, "pages": {}, "alias_index": {}, "sheet_dag": {}}
    pp.merge_into_index(index, s_entries, s_aliases, s_tree)

    if not no_pdf and pdfs:
        from ..config import pdf_page_cap
        cap = pdf_page_cap()                  # env > config.yaml > default (= PDF_PAGE_CAP)
        cap = cap if cap > 0 else None       # 0/negative → no cap (vision the whole PDF)
        used_docs = set(used_names)
        p_entries, p_aliases, p_tree, documents = ingest_pdfs(pdfs, root, pages_dir, used_docs, cap)
        pp.merge_into_index(index, p_entries, p_aliases, p_tree)
        index["documents"] = documents
        print(f"  ingested {len(p_entries)} PDF page(s) across {len(documents)} document(s)")

    dedup_alias_index(index, tuple(alias_stopwords()))

    out_json, out_md = wiki_index_json(), wiki_index_md()
    out_json.parent.mkdir(parents=True, exist_ok=True)
    out_json.write_text(json.dumps(index, ensure_ascii=False, indent=2), encoding="utf-8")
    out_md.write_text(render_md(index), encoding="utf-8")
    print(f"==> wrote {out_json}  (pages={len(index['pages'])}, aliases={len(index['alias_index'])})")
    return index


def plan_report(root: Path, only: str | None = None) -> dict[str, Counter]:
    """Offline curation dry-run: what gets ingested, by section / extension / grain. No I/O writes."""
    exclude, include = load_policy()
    files = curate(root, exclude, include)
    if only:
        only = _nfc(only)
        files = [f for f in files if section_of(f, root).startswith(only)]
    by_section: Counter = Counter(section_of(f, root) for f in files)
    by_ext: Counter = Counter(f.suffix.lower() for f in files)
    sheets, pdfs, unconverted = resolve_grains(files)
    by_grain = Counter({"spreadsheets": len(sheets), "pdfs": len(pdfs),
                        "unconverted (need convert)": len(unconverted)})
    print(f"==> curate dry-run over {root}  ({len(files)} file(s) kept)")
    for title, ctr in (("by section", by_section), ("by extension", by_ext), ("by grain", by_grain)):
        print(f"\n  {title}:")
        for k, v in ctr.most_common():
            print(f"    {v:>4}  {k}")
    return {"section": by_section, "ext": by_ext, "grain": by_grain}


def main(argv: list[str]) -> int:
    ap = argparse.ArgumentParser(description="Ingest a nested document data room into a wiki.")
    ap.add_argument("root", nargs="?", default=None, help="corpus root (default: config convert.root)")
    ap.add_argument("--plan", action="store_true", help="curation dry-run + counts (no build)")
    ap.add_argument("--only", default=None, help="restrict to one top section, e.g. '3.' (pilot)")
    ap.add_argument("--convert", action="store_true", help="normalize legacy formats first (convert.py)")
    ap.add_argument("--no-pdf", action="store_true", help="spreadsheets only (skip the vision stage)")
    args = ap.parse_args(argv)

    root = Path(args.root) if args.root else convert_root()
    if not root.exists():
        print(f"!! corpus root not found: {root}")
        return 1

    if args.plan:
        plan_report(root, only=args.only)
        return 0

    if args.convert:
        from ..convert import CONVERSIONS, convert, plan as convert_plan
        print("==> normalizing legacy formats (convert.py) ...")
        convert(convert_plan(root, CONVERSIONS))

    build(root, only=args.only, no_pdf=args.no_pdf)
    return 0


if __name__ == "__main__":
    # Smoke: a curation dry-run over the configured corpus root (offline — no LLM, no soffice).
    raise SystemExit(main(sys.argv[1:] or ["--plan"]))
