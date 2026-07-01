"""Data-room ingest logic — deterministic + offline (no vLLM, no LibreOffice).

Pins curation (exclude/include globs), grain resolution (legacy → converted sibling), page-name
namespacing/collision, the spreadsheet grid→alias extraction, and a full ``--no-pdf`` build over a
tiny real workbook written to a tmp tree. The PDF/vision stream (the only LLM path) is not exercised
here; ``build(..., no_pdf=True)`` covers the assembly.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from src.stella_kb.wiki import data_room as dr


def _corpus(tmp_path: Path) -> Path:
    """A miniature data room mirroring v0.3's shape: one kept ledger, plus files the four bulk
    exclude patterns must drop and one include-readmitted 정관."""
    root = tmp_path / "data"
    files = [
        "2. 재무/2.6. 투자자산/2.6.3. 수익증권/summary.xlsx",          # kept
        "5. 법무,인사/5.4. 영업계약/policy_약관.pdf",                  # excluded (영업계약)
        "2. 재무/2.6. 투자자산/2.6.3.1. 수익증권 평가보고서_231231/r1.pdf",  # excluded (per-security)
        "4. 세무/4.4. 부가가치세/vat_2024.xlsx",                       # excluded (bulk tax)
        "1. 회사일반현황/1.3. 정관 및 사규/(1-1800) 임원성과연봉규정.pdf",  # excluded (bylaws)
        "1. 회사일반현황/1.3. 정관 및 사규/(1-100) 정관 20250331.pdf",  # excluded then re-admitted
    ]
    for rel in files:
        p = root / rel
        p.parent.mkdir(parents=True, exist_ok=True)
        p.write_bytes(b"")
    return root


def test_curate_excludes_bulk_and_readmits_include(tmp_path):
    # The exclude/include *mechanism* is tested with explicit globs (not DEFAULT_EXCLUDE, which is
    # now full-ingest — see test_default_exclude_is_full_ingest). exclude drops a bulk set; include
    # rescues one file from that drop.
    root = _corpus(tmp_path)
    exclude = ["*5.4. 영업계약*", "*2.6.3.*평가보고서*", "*4.4. 부가가치세*", "*1.3. 정관 및 사규*"]
    include = ["*1.3. 정관 및 사규*정관*"]
    kept = {p.relative_to(root).as_posix() for p in dr.curate(root, exclude, include)}
    assert any("2.6.3. 수익증권/summary.xlsx" in k for k in kept)        # core kept
    assert not any("영업계약" in k for k in kept)                          # 약관 dropped
    assert not any("평가보고서" in k for k in kept)                        # per-security dropped
    assert not any("부가가치세" in k for k in kept)                        # bulk tax dropped
    assert not any("임원성과연봉규정" in k for k in kept)                  # bylaws dropped
    assert any("정관 20250331" in k for k in kept)                         # but 정관 re-admitted


def test_default_exclude_is_full_ingest(tmp_path):
    # As of the full re-admission, DEFAULT_EXCLUDE drops *only* the two unopenable security-locked
    # files — every DD-relevant file (incl. the former boilerplate bulk-drops) is kept.
    root = _corpus(tmp_path)
    (root / "2. 재무/2.3. 자산/유형자산, 투자부동산 연도별 세부명세서 (FY2023.4분기).xlsx").parent.mkdir(
        parents=True, exist_ok=True)
    (root / "2. 재무/2.3. 자산/유형자산, 투자부동산 연도별 세부명세서 (FY2023.4분기).xlsx").write_bytes(b"")
    kept = {p.relative_to(root).as_posix() for p in dr.curate(root, dr.DEFAULT_EXCLUDE, [])}
    assert any("영업계약" in k for k in kept)        # boilerplate now KEPT (full ingest)
    assert any("평가보고서" in k for k in kept)
    assert any("부가가치세" in k for k in kept)
    assert not any("FY2023.4분기" in k for k in kept)  # but the security-locked file stays dropped


def test_load_policy_falls_back_to_default(monkeypatch, tmp_path):
    monkeypatch.setenv("MNA_WIKI_CURATE", str(tmp_path / "nope.yaml"))
    exclude, include = dr.load_policy()
    assert exclude == dr.DEFAULT_EXCLUDE and include == []


def test_resolve_grains_follows_converted_sibling(tmp_path):
    root = tmp_path / "d"
    (root / "f").mkdir(parents=True)
    xlsx = root / "f" / "a.xlsx"; xlsx.write_bytes(b"")
    native_pdf = root / "f" / "b.pdf"; native_pdf.write_bytes(b"")
    doc = root / "f" / "c.doc"; doc.write_bytes(b"")
    (root / "f" / "c.pdf").write_bytes(b"")                 # c.doc's converted sibling
    orphan = root / "f" / "d.docx"; orphan.write_bytes(b"")  # no sibling → unconverted
    sheets, pdfs, unconverted = dr.resolve_grains([xlsx, native_pdf, doc, orphan])
    assert sheets == [xlsx]
    assert set(pdfs) == {native_pdf, root / "f" / "c.pdf"}   # native + converted, deduped
    assert unconverted == [orphan]


def test_resolve_grains_dedups_converted_xls(tmp_path):
    # convert.py runs without --replace, so a legacy a.xls AND its converted a.xlsx both survive on
    # disk and both land in the curated list. resolve_grains must yield the .xlsx ONCE — the .xls
    # branch (follows the sibling) and the .xlsx branch must not both append it (regression: dup page).
    root = tmp_path / "d"
    (root / "f").mkdir(parents=True)
    xls = root / "f" / "a.xls"; xls.write_bytes(b"")
    xlsx = root / "f" / "a.xlsx"; xlsx.write_bytes(b"")      # convert.py's non-destructive sibling
    sheets, pdfs, unconverted = dr.resolve_grains([xls, xlsx])
    assert sheets == [xlsx]                                   # exactly one, not [xlsx, xlsx]
    assert (pdfs, unconverted) == ([], [])


def test_sheet_page_name_namespaces_and_uniquifies(tmp_path):
    root = tmp_path / "d"
    p = root / "2. 재무" / "2.6.3. 수익증권" / "x.xlsx"
    p.parent.mkdir(parents=True)
    used: set[str] = set()
    n1 = dr.sheet_page_name(p, "Sheet1", root, 1, used)       # single sheet → no __sheet suffix
    n2 = dr.sheet_page_name(p, "Sheet1", root, 1, used)       # same again → uniquified
    assert n1 == "2.6.3. 수익증권__x"
    assert n2 == "2.6.3. 수익증권__x #2"
    multi = dr.sheet_page_name(p, "PL", root, 3, set())       # >1 sheet → __sheet suffix
    assert multi == "2.6.3. 수익증권__x__PL"


def test_dump_sheet_grid_and_aliases(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "관리수수료"; ws["B1"] = 1234.5     # label + number
    ws["A2"] = "AUM";       ws["B2"] = "55%"      # second label + a numeric-ish (dropped)
    md, pairs = dr.dump_sheet("p", ws)
    assert ("관리수수료", "A1") in pairs
    assert ("AUM", "A2") in pairs
    assert ("55%", "B2") not in pairs              # numeric-ish text dropped
    assert "| **1** |" in md and "1234.5" in md    # grid rendered with cached values + row refs
    assert "non-empty rows" not in md              # the redundant grid-summary line is gone (LLM noise)
    assert "×" not in md                           # no "N rows × M cols" boilerplate


def test_build_no_pdf_assembles_servable_index(monkeypatch, tmp_path):
    monkeypatch.setenv("MNA_WIKI_DATA", str(tmp_path / "out"))
    monkeypatch.setenv("MNA_WIKI_CURATE", str(tmp_path / "absent.yaml"))   # use DEFAULT_EXCLUDE
    root = tmp_path / "data"
    led = root / "2. 재무" / "2.6.3. 수익증권" / "fees.xlsx"
    led.parent.mkdir(parents=True)
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "관리보수"
    ws["A1"] = "항목"; ws["B1"] = "값"
    ws["A2"] = "관리수수료"; ws["B2"] = 1000
    wb.save(led)

    from src.stella_kb.wiki.index import render_md   # must not raise on the assembled records
    index = dr.build(root, no_pdf=True)

    assert len(index["pages"]) == 1
    (name,) = list(index["pages"])
    page = index["pages"][name]
    assert page["section"] == "2. 재무" and page["group"] == "2.6.3. 수익증권"
    assert {"kind", "case", "unit", "period", "depends_on", "feeds_into"} <= set(page)  # render_md needs these
    assert dr.pp._norm("관리수수료") in index["alias_index"]


def test_dump_spreadsheets_skips_chartsheet(monkeypatch, tmp_path):
    # A real data-room workbook can carry a Chartsheet (chart-only tab): it loads fine in read_only
    # and shows up in ``sheetnames``, but the object has no ``iter_rows`` — calling it crashed the
    # whole build. Faithful stand-in (openpyxl's own create_chartsheet round-trips differently): a
    # workbook whose 'chart' sheet lacks iter_rows. dump must skip it and still emit the data sheet.
    root = tmp_path / "data"
    xlsx = root / "2. 재무" / "2.6.3. 수익증권" / "withchart.xlsx"
    xlsx.parent.mkdir(parents=True)
    xlsx.write_bytes(b"")                         # content irrelevant — load_workbook is mocked

    class _Chart:                                 # Chartsheet stand-in: no iter_rows
        pass

    class _WS:
        def iter_rows(self, values_only=True):
            yield ("관리수수료", 1000)

    class _WB:
        sheetnames = ["data", "chart"]
        def __getitem__(self, k):
            return _WS() if k == "data" else _Chart()
        def close(self):
            pass

    monkeypatch.setattr(dr.openpyxl, "load_workbook", lambda *a, **k: _WB())
    entries, aliases, _tree = dr.dump_spreadsheets([xlsx], root, tmp_path / "pages", set())
    assert len(entries) == 1                       # only 'data' → one page; chartsheet skipped, no crash
    assert dr.pp._norm("관리수수료") in aliases


def test_pdf_page_cap_accessor_env_override(monkeypatch):
    # the cap goes through config (env > yaml > default), not a bare os.getenv in data_room
    from src.stella_kb import config
    monkeypatch.delenv("MNA_PDF_PAGE_CAP", raising=False)
    assert config.pdf_page_cap() == dr.PDF_PAGE_CAP          # default mirrors the module constant
    monkeypatch.setenv("MNA_PDF_PAGE_CAP", "5")
    assert config.pdf_page_cap() == 5                        # env override wins, typed as int
